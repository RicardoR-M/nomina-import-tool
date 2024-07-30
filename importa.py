from datetime import datetime
from fnmatch import fnmatch
from time import time
from warnings import simplefilter

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.datetime import from_excel
from dotenv import load_dotenv

# noinspection PyUnresolvedReferences
from config.adm import adm, adm_db, adm_db_table, adm_xl_firstDataRow, adm_xl_firstDataCol, adm_xl_colNameRow, adm_maxColNumber, adm_xl_hoja, adm_xl_colCheckFlag
# noinspection PyUnresolvedReferences
from config.tec import tec, tec_db, tec_db_table, tec_xl_firstDataRow, tec_xl_firstDataCol, tec_xl_colNameRow, tec_maxColNumber, tec_xl_hoja, tec_xl_colCheckFlag
# noinspection PyUnresolvedReferences
from config.post import post, post_db, post_db_table, post_xl_firstDataRow, post_xl_firstDataCol, post_xl_colNameRow, post_maxColNumber, post_xl_hoja, post_xl_colCheckFlag
# noinspection PyUnresolvedReferences
from config.at import at, at_db, at_db_table, at_xl_firstDataRow, at_xl_firstDataCol, at_xl_colNameRow, at_maxColNumber, at_xl_hoja, at_xl_colCheckFlag
# noinspection PyUnresolvedReferences
from config.mesa_val import mval, mval_db, mval_db_table, mval_xl_firstDataRow, mval_xl_firstDataCol, mval_xl_colNameRow, mval_maxColNumber, mval_xl_hoja, mval_xl_colCheckFlag
# noinspection PyUnresolvedReferences
from config.cont_seg import segcont, segcont_db, segcont_db_table, segcont_xl_firstDataRow, segcont_xl_firstDataCol, segcont_xl_colNameRow, segcont_maxColNumber, segcont_xl_hoja, segcont_xl_colCheckFlag
from dbconfig import save_to_db

def importa_informe(archivo_cz, columnas_cz, db_name, db_table, first_data_row, first_data_col, col_name_row, max_col, hoja, col_check_flag, semana, mes, anho):
    time_start = time()
    simplefilter("ignore")  # Desactiva el warning de graficos cuando se utiliza load_workbook
    wb = load_workbook(filename=archivo_cz, data_only=True, keep_vba=False, keep_links=False, read_only=True)

    if hoja not in wb.sheetnames:
        print(f"No se encontro la hoja {hoja}")
        quit()

    ws = wb[hoja]

    for row in ws.iter_rows(min_row=col_name_row, max_row=col_name_row, max_col=max_col):  # , values_only=True):
        if row[column_index_from_string(col_check_flag) - 1] is None:  # iterrows empieza desde 0 por eso se resta 1 cuando se combierte la letra de la col a número
            break

        # Ubica la letra de columna donde se encuentra la cabesera registrada
        for cabecera in columnas_cz:
            flag_encontro = False
            if cabecera['estado'] == 'fijo':
                valor = cabecera['colName']
            else:
                valor = f'{cabecera["colName"]} *{semana}*'

            for col in range(column_index_from_string(first_data_col) - 1, max_col):
                if row[col].value is None:
                    break
                if fnmatch(row[col].value, valor):
                    print(f'{valor} - {row[col].column_letter}')
                    cabecera['col'] = row[col].column_letter
                    flag_encontro = True
                    break
            if not flag_encontro:
                raise ValueError(f'No se encontró la columna: {cabecera["colName"]}')

    lista_evas = []
    for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row, max_col=max_col, values_only=True):
        dic_temp = {}
        if row[column_index_from_string(col_check_flag) - 1] is None:  # iterrows empieza desde 0 por eso se resta 1 cuando se combierte la letra de la col a número
            break

        for col in columnas_cz:
            valor = row[column_index_from_string(col['col']) - 1]
            if col['tipo'] == 'str':
                valor = str(valor)
            elif col['tipo'] == 'cf':
                valor = valor / 100.0
            elif col['tipo'] == 'hora':
                valor = valor.strftime("%H:%M:%S.%f")
            elif col['tipo'] == 'trim_str':
                if valor is not None:
                    valor = str(valor).strip()
            elif col['tipo'] == 'xl_date':
                if valor is not None:
                    if valor < 4000:
                        valor = None
                    else:
                        valor = from_excel(valor)
            elif col['tipo'] == 'str_to_datetime':
                if valor is not None:
                    valor = datetime.strptime(valor, '%d/%m/%Y %H:%M:%S')
            elif col['tipo'] == 'dni':
                if valor is not None:
                    valor = str(valor).upper().replace('C.E', '').strip()
                    if len(valor) < 8:
                        valor = '0' * (8 - len(valor)) + valor

            if valor == '-':
                valor = None
            if isinstance(valor, str):  # Si valor es string le hcace strip antes de agregarlo a la lista temporal
                valor = valor.strip()

            dic_temp[col['sqlName']] = valor
            dic_temp['SEMANA'] = semana
            dic_temp['AÑO'] = anho
            dic_temp['MES'] = mes
            dic_temp['N_MES'] = n_to_mes(mes)
        lista_evas.append(dic_temp)

    wb.close()
    save_to_db(db_name, db_table, lista_evas, semana, mes, anho)
    simplefilter("default")  # Habilita los warnings nuevamente
    time_end = time()
    print(f"TIEMPO: {time_end - time_start}")


def n_to_mes(n):
    if n == 1:
        return 'ENERO'
    if n == 2:
        return 'FEBRERO'
    if n == 3:
        return 'MARZO'
    if n == 4:
        return 'ABRIL'
    if n == 5:
        return 'MAYO'
    if n == 6:
        return 'JUNIO'
    if n == 7:
        return 'JULIO'
    if n == 8:
        return 'AGOSTO'
    if n == 9:
        return 'SEPTIEMBRE'
    if n == 10:
        return 'OCTUBRE'
    if n == 11:
        return 'NOVIEMBRE'
    if n == 12:
        return 'DICIEMBRE'


if __name__ == '__main__':
    load_dotenv()

    ruta_informe = './nominas/tec.xlsx'
    importa_informe(ruta_informe, tec, tec_db, tec_db_table, tec_xl_firstDataRow, tec_xl_firstDataCol, tec_xl_colNameRow, tec_maxColNumber, tec_xl_hoja, tec_xl_colCheckFlag, 31, 7, 2024)
    ruta_informe = './nominas/adm.xlsx'
    importa_informe(ruta_informe, adm, adm_db, adm_db_table, adm_xl_firstDataRow, adm_xl_firstDataCol, adm_xl_colNameRow, adm_maxColNumber, adm_xl_hoja, adm_xl_colCheckFlag, 31, 7, 2024)
    ruta_informe = './nominas/post.xlsx'
    importa_informe(ruta_informe, post, post_db, post_db_table, post_xl_firstDataRow, post_xl_firstDataCol, post_xl_colNameRow, post_maxColNumber, post_xl_hoja, post_xl_colCheckFlag, 31, 7, 2024)
    ruta_informe = './nominas/at.xlsx'
    importa_informe(ruta_informe, at, at_db, at_db_table, at_xl_firstDataRow, at_xl_firstDataCol, at_xl_colNameRow, at_maxColNumber, at_xl_hoja, at_xl_colCheckFlag, 31, 7, 2024)
    ruta_informe = './nominas/mesa.xlsx'
    importa_informe(ruta_informe, mval, mval_db, mval_db_table, mval_xl_firstDataRow, mval_xl_firstDataCol, mval_xl_colNameRow, mval_maxColNumber, mval_xl_hoja, mval_xl_colCheckFlag, 31, 7, 2024)
    ruta_informe = './nominas/contseg.xlsx'
    importa_informe(ruta_informe, segcont, segcont_db, segcont_db_table, segcont_xl_firstDataRow, segcont_xl_firstDataCol, segcont_xl_colNameRow, segcont_maxColNumber, segcont_xl_hoja, segcont_xl_colCheckFlag, 31, 7, 2024)

